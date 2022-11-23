#! /usr/bin/env python3
# blankRowInserter.py
import sys
import openpyxl

if len(sys.argv) < 4:
    print('Usage: ./blankRowInserter.py <n> <m> <filename>.xlsx where <n> and <m> are integers')
    sys.exit()

if not sys.argv[1].isdigit() or not sys.argv[2].isdigit():
    print('./blankRowInserter.py <n> <m> <filename>.xlsx where <n> and <m> must be integers')
    sys.exit()

n = int(sys.argv[1])
m = int(sys.argv[2])
filename = sys.argv[3]
wb = openpyxl.load_workbook(filename)
sheet = wb.active

before_n = []
after_n = []
max_col = sheet.max_column
max_row = sheet.max_row

for i in range(1, n):
    for j in range(1, max_col+1):
        before_n.append(sheet.cell(row=i, column=j).value)

for i in range(n, max_row+1):
    for j in range(1, max_col+1):
        after_n.append(sheet.cell(row=i, column=j).value)

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for i in range(1, n):
    for j in range(1, max_col+1):
        new_sheet.cell(row=i, column=j).value = before_n.pop(0)

for i in range(n+m+1, max_row+1):
    for j in range(1, max_col+1):
        new_sheet.cell(row=i, column=j).value = after_n.pop(0)

new_wb.save('edited_' + filename)


### Alternatively,
# for i in range(1, max_row+1):
#     for j in range(1, max_col+1):
#         if i < n:
#             new_sheet.cell(row=i, column=j).value = sheet.cell(row=i, column=j).value
#         else:
#             new_sheet.cell(row=i+m, column=j).value = sheet.cell(row=i, column=j).value
#
# new_wb.save('edited_' + filename)
